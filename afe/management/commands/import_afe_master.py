"""
Import AFE master data.

1. Seed `AFETemplate` with the 58 static F.CONT lines (hard-coded — stable).
2. Import `RateCardItem` rows from sheet `D.MATE` of the reference workbook.

Usage:
    python manage.py import_afe_master "path/to/AFE_v.2017_Kontrak JTB_Update April 2026.xlsx"

Flags:
    --templates-only   Only seed AFETemplate catalog, skip D.MATE import.
    --rates-only       Only import D.MATE rate card (assumes templates exist).
    --dry-run          Parse + report, rollback at the end.
"""
from __future__ import annotations

from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from afe.models import (
    AFECategory,
    AFESection,
    AFETemplate,
    CalcMethod,
    PhaseFlag,
    RateCardItem,
)


# ---------------------------------------------------------------------------
# 58-line F.CONT catalog (pure Python, no Excel parsing)
# ---------------------------------------------------------------------------
# (line_code, name, category, section, calc_method, is_subtotal_row)
TEMPLATE_ROWS: list[tuple[str, str, str, str, str, bool]] = [
    # ---- TANGIBLE COST ------------------------------------------------
    ("1",  "TANGIBLE COST",                       AFECategory.TANGIBLE,   AFESection.TANGIBLE,   CalcMethod.MANUAL,            True),
    ("2",  "Casing",                              AFECategory.TANGIBLE,   AFESection.TANGIBLE,   CalcMethod.PER_CASING_WEIGHT, False),
    ("3",  "Casing Accessories",                  AFECategory.TANGIBLE,   AFESection.TANGIBLE,   CalcMethod.LUMP_SUM,          False),
    ("4",  "Tubing",                              AFECategory.TANGIBLE,   AFESection.TANGIBLE,   CalcMethod.PER_CASING_WEIGHT, False),
    ("5",  "Well Equipment - Surface",            AFECategory.TANGIBLE,   AFESection.TANGIBLE,   CalcMethod.LUMP_SUM,          False),
    ("6",  "Well Equipment - Subsurface",         AFECategory.TANGIBLE,   AFESection.TANGIBLE,   CalcMethod.LUMP_SUM,          False),
    ("7",  "Other Tangible Cost",                 AFECategory.TANGIBLE,   AFESection.TANGIBLE,   CalcMethod.LUMP_SUM,          False),
    ("8",  "TOTAL TANGIBLE COST",                 AFECategory.TANGIBLE,   AFESection.TANGIBLE,   CalcMethod.MANUAL,            True),

    # ---- INTANGIBLE — Preparation -------------------------------------
    ("9",  "INTANGIBLE COST",                     AFECategory.INTANGIBLE, AFESection.PREPARATION, CalcMethod.MANUAL,            True),
    ("10", "PREPARATION AND TERMINATION",         AFECategory.INTANGIBLE, AFESection.PREPARATION, CalcMethod.MANUAL,            True),
    ("11", "Surveys",                             AFECategory.INTANGIBLE, AFESection.PREPARATION, CalcMethod.PER_METER_DEPTH,   False),
    ("12", "Location Staking and Positioning",    AFECategory.INTANGIBLE, AFESection.PREPARATION, CalcMethod.LUMP_SUM,          False),
    ("13", "Wellsite and Access Road Preparation", AFECategory.INTANGIBLE, AFESection.PREPARATION, CalcMethod.LUMP_SUM,          False),
    ("14", "Service Lines & Communications",      AFECategory.INTANGIBLE, AFESection.PREPARATION, CalcMethod.LUMP_SUM,          False),
    ("15", "Water Systems",                       AFECategory.INTANGIBLE, AFESection.PREPARATION, CalcMethod.LUMP_SUM,          False),
    ("16", "Rigging Up / Rigging Down",           AFECategory.INTANGIBLE, AFESection.PREPARATION, CalcMethod.LUMP_SUM,          False),
    ("17", "SUBTOTAL PREPARATION",                AFECategory.INTANGIBLE, AFESection.PREPARATION, CalcMethod.MANUAL,            True),

    # ---- INTANGIBLE — Drilling / Workover -----------------------------
    ("18", "DRILLING / WORKOVER / WELL SERVICE OPERATION", AFECategory.INTANGIBLE, AFESection.DRILLING, CalcMethod.MANUAL, True),
    ("19", "Contract Rig",                        AFECategory.INTANGIBLE, AFESection.DRILLING,    CalcMethod.RIG_DAYS_RATE,     False),
    ("20", "Drilling Rig Crew / Contract Rig Crew", AFECategory.INTANGIBLE, AFESection.DRILLING,  CalcMethod.RIG_DAYS_RATE,     False),
    ("21", "Mud, Chemical & Engineering Services", AFECategory.INTANGIBLE, AFESection.DRILLING,   CalcMethod.DHB_CB_SPLIT,      False),
    ("22", "Water",                               AFECategory.INTANGIBLE, AFESection.DRILLING,    CalcMethod.RIG_DAYS_RATE,     False),
    ("23", "Bits, Reamer and Core Heads",         AFECategory.INTANGIBLE, AFESection.DRILLING,    CalcMethod.LUMP_SUM,          False),
    ("24", "Equipment Rent",                      AFECategory.INTANGIBLE, AFESection.DRILLING,    CalcMethod.RIG_DAYS_RATE,     False),
    ("25", "Directional Drilling and Survey",     AFECategory.INTANGIBLE, AFESection.DRILLING,    CalcMethod.LUMP_SUM,          False),
    ("26", "Diving Services",                     AFECategory.INTANGIBLE, AFESection.DRILLING,    CalcMethod.LUMP_SUM,          False),
    ("27", "Casing Installation",                 AFECategory.INTANGIBLE, AFESection.DRILLING,    CalcMethod.LUMP_SUM,          False),
    ("28", "Cement, Cementing and Pump Fees",     AFECategory.INTANGIBLE, AFESection.DRILLING,    CalcMethod.DHB_CB_SPLIT,      False),
    ("29", "SUBTOTAL DRILLING",                   AFECategory.INTANGIBLE, AFESection.DRILLING,    CalcMethod.MANUAL,            True),

    # ---- INTANGIBLE — Formation Evaluation ----------------------------
    ("30", "FORMATION EVALUATION",                AFECategory.INTANGIBLE, AFESection.FORMATION,   CalcMethod.MANUAL,            True),
    ("31", "Coring",                              AFECategory.INTANGIBLE, AFESection.FORMATION,   CalcMethod.LUMP_SUM,          False),
    ("32", "Mud Logging Services",                AFECategory.INTANGIBLE, AFESection.FORMATION,   CalcMethod.PER_METER_DEPTH,   False),
    ("33", "Drill Stem Test",                     AFECategory.INTANGIBLE, AFESection.FORMATION,   CalcMethod.LUMP_SUM,          False),
    ("34", "Open Hole Electrical Logging Services", AFECategory.INTANGIBLE, AFESection.FORMATION, CalcMethod.LUMP_SUM,          False),
    ("35", "SUBTOTAL FORMATION",                  AFECategory.INTANGIBLE, AFESection.FORMATION,   CalcMethod.MANUAL,            True),

    # ---- INTANGIBLE — Completion --------------------------------------
    ("36", "COMPLETION",                          AFECategory.INTANGIBLE, AFESection.COMPLETION,  CalcMethod.MANUAL,            True),
    ("37", "Casing Liner and Tubing Installation", AFECategory.INTANGIBLE, AFESection.COMPLETION, CalcMethod.LUMP_SUM,          False),
    ("38", "Cement, Cementing and Pump Fees (Completion)", AFECategory.INTANGIBLE, AFESection.COMPLETION, CalcMethod.LUMP_SUM, False),
    ("39", "Cased Hole Electrical Logging Services", AFECategory.INTANGIBLE, AFESection.COMPLETION, CalcMethod.LUMP_SUM,        False),
    ("40", "Perforating and Wireline Services",   AFECategory.INTANGIBLE, AFESection.COMPLETION,  CalcMethod.LUMP_SUM,          False),
    ("41", "Stimulation Treatment",               AFECategory.INTANGIBLE, AFESection.COMPLETION,  CalcMethod.LUMP_SUM,          False),
    ("42", "Production Test",                     AFECategory.INTANGIBLE, AFESection.COMPLETION,  CalcMethod.LUMP_SUM,          False),
    ("43", "SUBTOTAL COMPLETION",                 AFECategory.INTANGIBLE, AFESection.COMPLETION,  CalcMethod.MANUAL,            True),

    # ---- INTANGIBLE — General / Overhead ------------------------------
    ("44", "GENERAL",                             AFECategory.INTANGIBLE, AFESection.GENERAL,     CalcMethod.MANUAL,            True),
    ("45", "Project Management Team",             AFECategory.INTANGIBLE, AFESection.GENERAL,     CalcMethod.RIG_DAYS_RATE,     False),
    ("46", "Insurance",                           AFECategory.INTANGIBLE, AFESection.GENERAL,     CalcMethod.LUMP_SUM,          False),
    ("47", "Permits and Fees",                    AFECategory.INTANGIBLE, AFESection.GENERAL,     CalcMethod.LUMP_SUM,          False),
    ("48", "Marine Rental and Charters",          AFECategory.INTANGIBLE, AFESection.GENERAL,     CalcMethod.LUMP_SUM,          False),
    ("49", "Helicopter Aviation and Charges",     AFECategory.INTANGIBLE, AFESection.GENERAL,     CalcMethod.LUMP_SUM,          False),
    ("50", "Land Transportation",                 AFECategory.INTANGIBLE, AFESection.GENERAL,     CalcMethod.RIG_DAYS_RATE,     False),
    ("51", "Other Transportation",                AFECategory.INTANGIBLE, AFESection.GENERAL,     CalcMethod.LUMP_SUM,          False),
    ("52", "Fuel and Lubricants",                 AFECategory.INTANGIBLE, AFESection.GENERAL,     CalcMethod.RIG_DAYS_RATE,     False),
    ("53", "Camp Facilities",                     AFECategory.INTANGIBLE, AFESection.GENERAL,     CalcMethod.RIG_DAYS_RATE,     False),
    ("54", "Allocated Overhead Field Office",     AFECategory.INTANGIBLE, AFESection.GENERAL,     CalcMethod.LUMP_SUM,          False),
    ("55", "Allocated Overhead Jakarta Office",   AFECategory.INTANGIBLE, AFESection.GENERAL,     CalcMethod.LUMP_SUM,          False),
    ("56", "SUBTOTAL GENERAL",                    AFECategory.INTANGIBLE, AFESection.GENERAL,     CalcMethod.MANUAL,            True),
    ("57", "TOTAL INTANGIBLE COST",               AFECategory.INTANGIBLE, AFESection.GENERAL,     CalcMethod.MANUAL,            True),
    ("58", "TOTAL COSTS",                         AFECategory.INTANGIBLE, AFESection.GENERAL,     CalcMethod.MANUAL,            True),
]


def _to_decimal(value) -> Decimal | None:
    if value in (None, ""):
        return None
    try:
        return Decimal(str(value).strip())
    except (InvalidOperation, ValueError):
        return None


def _text(value) -> str:
    return str(value).strip() if value is not None else ""


class Command(BaseCommand):
    help = "Seed AFE 58-line template catalog and import rate card from D.MATE."

    def add_arguments(self, parser):
        parser.add_argument("excel_path", nargs="?", type=str,
                            help="Path to AFE_*.xlsx (optional if --templates-only)")
        parser.add_argument("--templates-only", action="store_true")
        parser.add_argument("--rates-only", action="store_true")
        parser.add_argument("--dry-run", action="store_true")

    def handle(self, *args, **options):
        do_templates = not options["rates_only"]
        do_rates = not options["templates_only"]

        if do_rates and not options["excel_path"]:
            raise CommandError("excel_path required (or pass --templates-only).")

        with transaction.atomic():
            if do_templates:
                self._seed_templates()
                self._seed_fallback_rates()

            if do_rates:
                path = Path(options["excel_path"])
                if not path.exists():
                    raise CommandError(f"File not found: {path}")
                try:
                    import openpyxl  # noqa: F401
                except ImportError as exc:
                    raise CommandError(
                        "openpyxl not installed. Run: pip install -r requirements.txt"
                    ) from exc

                from openpyxl import load_workbook
                self.stdout.write(f"Loading workbook {path.name} …")
                wb = load_workbook(path, data_only=True, read_only=True)
                self._import_dmate(wb)

            if options["dry_run"]:
                transaction.set_rollback(True)
                self.stdout.write(self.style.WARNING("Dry run — rolled back."))
            else:
                self.stdout.write(self.style.SUCCESS("Import complete."))

    # ------------------------------------------------------------------
    # Template catalog
    # ------------------------------------------------------------------
    def _seed_templates(self):
        created = 0
        updated = 0
        for order, row in enumerate(TEMPLATE_ROWS, start=1):
            line_code, name, cat, section, calc, is_sub = row
            obj, was_created = AFETemplate.objects.update_or_create(
                line_code=line_code,
                defaults={
                    "name": name,
                    "category": cat,
                    "section": section,
                    "calc_method": calc,
                    "is_subtotal_row": is_sub,
                    "order": order,
                },
            )
            if was_created:
                created += 1
            else:
                updated += 1
        self.stdout.write(
            f"  AFETemplate: total={AFETemplate.objects.count()} "
            f"(new {created}, updated {updated})"
        )

    # ------------------------------------------------------------------
    # Fallback rate card so calc engine produces non-zero output even
    # when D.MATE is not available. Values approximate typical 2017
    # JTB contract rates referenced in sheet `C2-19_RIG` and elsewhere.
    # ------------------------------------------------------------------
    def _seed_fallback_rates(self):
        from datetime import date
        FALLBACKS = [
            # (line_code, code, desc, uom, price_usd, phase)
            ("3",  "FB-CSG-ACC",  "Casing accessories (bulk)",              "lump",   35000,  PhaseFlag.BOTH),
            ("5",  "FB-WH-SURF",  "Wellhead / surface equipment",           "set",    85000,  PhaseFlag.BOTH),
            ("6",  "FB-WH-SUB",   "Subsurface equipment",                   "set",    40000,  PhaseFlag.BOTH),
            ("7",  "FB-TNG-OTH",  "Other tangibles",                        "lump",   15000,  PhaseFlag.BOTH),
            ("11", "FB-SURVEY",   "Topographic + soil survey",              "m",         12,  PhaseFlag.DHB),
            ("12", "FB-STAKING",  "Location staking & positioning",         "lump",    8000,  PhaseFlag.DHB),
            ("13", "FB-ROAD",     "Wellsite & access road preparation",     "lump",  120000,  PhaseFlag.DHB),
            ("14", "FB-COMM",     "Service lines & communications",         "lump",   15000,  PhaseFlag.DHB),
            ("15", "FB-WATER",    "Water systems",                          "lump",   12000,  PhaseFlag.DHB),
            ("16", "FB-RIGUP",    "Rig up / rig down",                      "lump",   95000,  PhaseFlag.DHB),
            ("19", "FB-RIG",      "Contract rig daily rate",                "day",   124000,  PhaseFlag.BOTH),
            ("20", "FB-RIG-CREW", "Rig crew daily allowance",               "day",     7500,  PhaseFlag.BOTH),
            ("21", "FB-MUD-DHB",  "Mud, chemical & engineering (DHB)",      "day",     9500,  PhaseFlag.DHB),
            ("21", "FB-MUD-CB",   "Mud, chemical & engineering (CB)",       "day",     4800,  PhaseFlag.CB),
            ("22", "FB-WATER-D",  "Water during drilling",                  "day",      650,  PhaseFlag.BOTH),
            ("23", "FB-BITS",     "Bits, reamer, core heads (set)",         "lump",  180000,  PhaseFlag.DHB),
            ("24", "FB-EQRENT",   "Drilling equipment rent",                "day",     4500,  PhaseFlag.BOTH),
            ("25", "FB-DIRDRL",  "Directional drilling & survey",           "lump",  180000,  PhaseFlag.DHB),
            ("27", "FB-CSG-INS",  "Casing installation service",            "lump",   55000,  PhaseFlag.DHB),
            ("28", "FB-CMT-DHB",  "Cement & pump fees (DHB)",               "day",     5500,  PhaseFlag.DHB),
            ("28", "FB-CMT-CB",   "Cement & pump fees (CB)",                "day",     6200,  PhaseFlag.CB),
            ("31", "FB-CORING",   "Coring service",                         "lump",   45000,  PhaseFlag.DHB),
            ("32", "FB-MUDLOG",   "Mud logging services",                   "m",         18,  PhaseFlag.DHB),
            ("33", "FB-DST",      "Drill stem test",                        "lump",   85000,  PhaseFlag.DHB),
            ("34", "FB-OHL",      "Open hole electrical logging",           "lump",  220000,  PhaseFlag.DHB),
            ("37", "FB-CSG-LINER","Liner & tubing installation",            "lump",   45000,  PhaseFlag.CB),
            ("38", "FB-CMT-LNR",  "Completion cementing",                   "lump",   32000,  PhaseFlag.CB),
            ("39", "FB-CHL",      "Cased hole electrical logging",          "lump",   95000,  PhaseFlag.CB),
            ("40", "FB-PERF",     "Perforating & wireline services",        "lump",  120000,  PhaseFlag.CB),
            ("41", "FB-STIM",     "Stimulation / acidizing",                "lump",  150000,  PhaseFlag.CB),
            ("42", "FB-TEST",     "Production test",                        "lump",   75000,  PhaseFlag.CB),
            ("45", "FB-PMT",      "Project management team",                "day",     3200,  PhaseFlag.BOTH),
            ("46", "FB-INS",      "Insurance",                              "lump",   45000,  PhaseFlag.BOTH),
            ("47", "FB-PERMIT",   "Permits & regulatory fees",              "lump",   18000,  PhaseFlag.BOTH),
            ("49", "FB-AIR",      "Helicopter / aviation charges",          "lump",   12000,  PhaseFlag.BOTH),
            ("50", "FB-TRNS-LD",  "Land transportation",                    "day",      850,  PhaseFlag.BOTH),
            ("51", "FB-TRNS-OTH", "Other transportation",                   "lump",   20000,  PhaseFlag.BOTH),
            ("52", "FB-FUEL",     "Fuel & lubricants",                      "day",     1600,  PhaseFlag.BOTH),
            ("53", "FB-CAMP",     "Camp facilities",                        "day",     1200,  PhaseFlag.BOTH),
            ("54", "FB-OH-FLD",   "Allocated overhead field office",        "lump",   25000,  PhaseFlag.BOTH),
            ("55", "FB-OH-JKT",   "Allocated overhead Jakarta office",      "lump",   35000,  PhaseFlag.BOTH),
        ]
        templates_by_code = {t.line_code: t for t in AFETemplate.objects.all()}
        created = updated = 0
        for line_code, code, desc, uom, price, phase in FALLBACKS:
            tpl = templates_by_code.get(line_code)
            _, was_created = RateCardItem.objects.update_or_create(
                code=code,
                defaults={
                    "description": desc,
                    "unit_of_measure": uom,
                    "unit_price_usd": Decimal(str(price)),
                    "afe_line": tpl,
                    "phase_flag": phase,
                    "material_type": "",
                    "source_sheet": "FALLBACK",
                    "effective_from": date.today(),
                },
            )
            if was_created:
                created += 1
            else:
                updated += 1
        self.stdout.write(
            f"  Fallback rates: new {created}, updated {updated}"
        )

    # ------------------------------------------------------------------
    # D.MATE rate card
    # Column layout (0-indexed, discovered at header row 83):
    #   47 = order within section
    #   48 = AFE line number (integer)
    #   49 = AFE line description (e.g. "CASING", "CONTRACT RIG")
    #   53 = material description
    #   58 = qty
    #   60 = UoM
    #   61 = unit price USD
    #   63 = material type
    #   69 = DHB/CB flag
    # ------------------------------------------------------------------
    def _import_dmate(self, wb):
        if "D.MATE" not in wb.sheetnames:
            self.stdout.write(self.style.WARNING("  D.MATE sheet missing, skipping."))
            return
        ws = wb["D.MATE"]

        COL_AFE_LINE = 48
        COL_AFE_DESC = 49
        COL_MAT_DESC = 53
        COL_UOM = 60
        COL_PRICE = 61
        COL_MAT_TYPE = 63
        COL_PHASE = 69

        # Pre-index templates by line_code
        templates_by_code = {t.line_code: t for t in AFETemplate.objects.all()}

        def _cell(row, idx):
            return row[idx] if idx < len(row) else None

        created = 0
        updated = 0
        skipped = 0
        seen_codes: set[str] = set()

        for i, row in enumerate(ws.iter_rows(min_row=86, values_only=True), start=86):
            if not row:
                continue
            line_val = _cell(row, COL_AFE_LINE)
            material_desc = _text(_cell(row, COL_MAT_DESC))
            price = _to_decimal(_cell(row, COL_PRICE))

            if not material_desc or price is None:
                skipped += 1
                continue

            # AFE line may be int or string like "19"
            if isinstance(line_val, (int, float)):
                line_code = str(int(line_val))
            else:
                line_code = _text(line_val)
            template = templates_by_code.get(line_code)

            uom = _text(_cell(row, COL_UOM))[:30]
            mat_type = _text(_cell(row, COL_MAT_TYPE))[:30]
            phase_raw = _text(_cell(row, COL_PHASE)).upper()
            if "DHB" in phase_raw and "CB" not in phase_raw:
                phase = PhaseFlag.DHB
            elif "CB" in phase_raw and "DHB" not in phase_raw:
                phase = PhaseFlag.CB
            else:
                phase = PhaseFlag.BOTH

            # Unique code per row (stable across imports): "<line>-<row_num>"
            code = f"L{line_code or 'X'}-{i}"
            if code in seen_codes:
                continue
            seen_codes.add(code)

            obj, was_created = RateCardItem.objects.update_or_create(
                code=code,
                defaults={
                    "description": material_desc[:300],
                    "unit_of_measure": uom,
                    "unit_price_usd": price,
                    "afe_line": template,
                    "phase_flag": phase,
                    "material_type": mat_type,
                    "source_sheet": "D.MATE",
                },
            )
            if was_created:
                created += 1
            else:
                updated += 1

        self.stdout.write(
            f"  RateCardItem: total={RateCardItem.objects.count()} "
            f"(new {created}, updated {updated}, skipped {skipped})"
        )
