from decimal import Decimal, InvalidOperation

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.http import HttpResponseForbidden
from django.shortcuts import get_object_or_404, redirect, render
from django.views.decorators.http import require_http_methods

from proposals.models import Proposal, ProposalStatus

from .forms import (
    AFEApprovalActionForm,
    AFEHeaderForm,
    AFELineComponentFormSet,
    AFELineFormSet,
    RateCardItemForm,
    RateCardUploadForm,
)
from .models import AFE, AFECategory, AFELine, AFEStatus, AFETemplate, RateCardItem, RateCardImportLog
from .services import approval
from .services.calc import generate_afe_from_proposal, recalculate_afe


# ---------------------------------------------------------------------------
# Permission helpers
# ---------------------------------------------------------------------------
def _require_rate_card_access(user):
    """Only Admin and Management can manage rate cards."""
    if user.is_superuser or getattr(user, "is_admin_role", False) or getattr(user, "is_management", False):
        return None
    return HttpResponseForbidden("Hanya Admin dan Management yang dapat mengelola daftar harga.")


# ---------------------------------------------------------------------------
# Dashboard & Inbox
# ---------------------------------------------------------------------------
@login_required
def dashboard(request):
    status = request.GET.get("status") or ""
    q = request.GET.get("q") or ""
    qs = AFE.objects.select_related("proposal", "proposal__well", "created_by").all()

    if request.user.is_engineer:
        qs = qs.filter(created_by=request.user)

    if status:
        qs = qs.filter(status=status)
    if q:
        qs = qs.filter(
            Q(doc_number__icontains=q)
            | Q(proposal__well__name__icontains=q)
            | Q(title__icontains=q)
        )

    return render(request, "afe/dashboard.html", {
        "afes": qs[:200],
        "status_choices": AFEStatus.choices,
        "current_status": status,
        "q": q,
        "rate_card_count": RateCardItem.objects.count(),
        "can_manage_rates": (
            request.user.is_superuser
            or getattr(request.user, "is_admin_role", False)
            or getattr(request.user, "is_management", False)
        ),
    })


@login_required
def inbox(request):
    user = request.user
    qs = AFE.objects.select_related("proposal", "proposal__well", "created_by")
    if user.is_supervisor:
        qs = qs.filter(status=AFEStatus.SUBMITTED)
    elif user.is_management:
        qs = qs.filter(status=AFEStatus.UNDER_REVIEW)
    elif user.is_engineer:
        qs = qs.filter(created_by=user, status=AFEStatus.REVISION)
    else:
        qs = qs.none()
    return render(request, "afe/inbox.html", {"afes": qs})


# ---------------------------------------------------------------------------
# Create from approved proposal
# ---------------------------------------------------------------------------
@login_required
@require_http_methods(["POST", "GET"])
def afe_create(request, proposal_pk):
    proposal = get_object_or_404(Proposal, pk=proposal_pk)
    user = request.user
    if not (user.is_engineer or user.is_admin_role):
        return HttpResponseForbidden("Hanya Drilling Engineer yang dapat membuat AFE.")
    if user.is_engineer and proposal.created_by != user:
        return HttpResponseForbidden("Anda hanya dapat membuat AFE dari proposal milik Anda sendiri.")
    if proposal.status != ProposalStatus.APPROVED:
        messages.error(request, "AFE hanya bisa dibuat dari Proposal yang sudah APPROVED.")
        return redirect("proposals:detail", pk=proposal.pk)

    prev = AFE.objects.filter(proposal=proposal).order_by("-version").first()
    version = (prev.version + 1) if prev else 1

    afe = AFE.objects.create(
        proposal=proposal,
        version=version,
        created_by=user,
        title=f"AFE {proposal.well.name} v{version}",
    )
    generate_afe_from_proposal(afe)
    messages.success(request, f"AFE v{version} dibuat. Silakan review & edit.")
    return redirect("afe:edit", pk=afe.pk)


# ---------------------------------------------------------------------------
# Detail
# ---------------------------------------------------------------------------
@login_required
def afe_detail(request, pk):
    afe = get_object_or_404(
        AFE.objects.select_related("proposal", "proposal__well", "created_by"),
        pk=pk,
    )
    # IDOR protection: engineers can only see their own AFEs
    if request.user.is_engineer and afe.created_by != request.user:
        return HttpResponseForbidden("Anda tidak memiliki akses ke AFE ini.")
    lines = (afe.lines
             .select_related("template", "rate_card_item")
             .prefetch_related("components")
             .order_by("template__order", "template__line_code"))

    # Donut chart: breakdown by section (non-subtotal lines only)
    section_totals: dict[str, float] = {}
    for line in lines:
        if line.template.is_subtotal_row:
            continue
        key = line.template.get_section_display()
        section_totals[key] = section_totals.get(key, 0.0) + float(line.final_usd)

    chart_labels = list(section_totals.keys())
    chart_values = [round(v, 2) for v in section_totals.values()]

    return render(request, "afe/detail.html", {
        "afe": afe,
        "lines": lines,
        "chart_labels": chart_labels,
        "chart_values": chart_values,
        "approval_form": AFEApprovalActionForm(),
        "approval_logs": afe.approval_logs.select_related("actor").all(),
        "can_edit": afe.can_edit(request.user),
        "can_submit": afe.can_submit(request.user),
        "can_review": afe.can_review(request.user),
        "can_approve": afe.can_approve(request.user),
    })


# ---------------------------------------------------------------------------
# Edit AFE header + lines
# ---------------------------------------------------------------------------
@login_required
def afe_edit(request, pk):
    afe = get_object_or_404(AFE, pk=pk)
    if not afe.can_edit(request.user):
        return HttpResponseForbidden("AFE tidak dapat diedit pada status saat ini.")

    if request.method == "POST":
        header_form = AFEHeaderForm(request.POST, instance=afe)
        formset = AFELineFormSet(request.POST, instance=afe)
        if header_form.is_valid() and formset.is_valid():
            header_form.save()
            formset.save()
            recalculate_afe(afe)
            messages.success(request, "AFE tersimpan.")
            if "save_and_continue" in request.POST:
                return redirect("afe:edit", pk=pk)
            return redirect("afe:detail", pk=pk)
    else:
        header_form = AFEHeaderForm(instance=afe)
        formset = AFELineFormSet(instance=afe)

    # Pair each form with its underlying line for the template
    lines_by_id = {l.id: l for l in afe.lines.select_related("template").prefetch_related("components")}
    rows = []
    for form in formset.forms:
        line = lines_by_id.get(form.instance.pk)
        rows.append((form, line))

    return render(request, "afe/edit.html", {
        "afe": afe,
        "header_form": header_form,
        "formset": formset,
        "rows": rows,
    })


# ---------------------------------------------------------------------------
# Edit components per AFE line
# ---------------------------------------------------------------------------
@login_required
def afe_line_components(request, pk, line_pk):
    afe = get_object_or_404(AFE, pk=pk)
    line = get_object_or_404(AFELine.objects.select_related("template"), pk=line_pk, afe=afe)
    if not afe.can_edit(request.user):
        return HttpResponseForbidden("AFE tidak dapat diedit pada status saat ini.")

    if request.method == "POST":
        formset = AFELineComponentFormSet(request.POST, instance=line)
        if formset.is_valid():
            formset.save()
            # Recalculate line total from components
            comp_total = sum(c.total_usd for c in line.components.all())
            if comp_total > 0:
                line.calculated_usd = comp_total
                line.save(update_fields=["calculated_usd"])
            recalculate_afe(afe)
            messages.success(request, f"Components untuk '{line.template.name}' disimpan.")
            if "save_and_continue" in request.POST:
                return redirect("afe:line_components", pk=pk, line_pk=line_pk)
            return redirect("afe:edit", pk=pk)
    else:
        formset = AFELineComponentFormSet(instance=line)

    return render(request, "afe/edit_components.html", {
        "afe": afe,
        "line": line,
        "formset": formset,
    })


# ---------------------------------------------------------------------------
# Regenerate AFE from rate cards
# ---------------------------------------------------------------------------
@login_required
@require_http_methods(["POST"])
def afe_regenerate(request, pk):
    afe = get_object_or_404(AFE, pk=pk)
    if not afe.can_edit(request.user):
        return HttpResponseForbidden("AFE tidak dapat diubah.")
    generate_afe_from_proposal(afe, overwrite=True)
    messages.success(request, "AFE di-regenerate dari rate card (override dihapus).")
    return redirect("afe:edit", pk=pk)


# ---------------------------------------------------------------------------
# Approval actions
# ---------------------------------------------------------------------------
@login_required
@require_http_methods(["POST"])
def afe_action(request, pk):
    afe = get_object_or_404(AFE, pk=pk)
    action = request.POST.get("action")
    comment = request.POST.get("comment", "").strip()

    try:
        if action == "submit":
            approval.submit(afe, request.user, comment)
            messages.success(request, "AFE disubmit ke supervisor.")
        elif action == "forward":
            approval.forward(afe, request.user, comment)
            messages.success(request, "AFE diteruskan ke management.")
        elif action == "request_revision":
            approval.request_revision(afe, request.user, comment)
            messages.info(request, "Revisi diminta.")
        elif action == "approve":
            approval.approve(afe, request.user, comment)
            messages.success(request, "AFE disetujui.")
        elif action == "reject":
            approval.reject(afe, request.user, comment)
            messages.warning(request, "AFE ditolak.")
        else:
            messages.error(request, "Aksi tidak dikenal.")
    except approval.ApprovalError as exc:
        messages.error(request, str(exc))

    return redirect("afe:detail", pk=pk)


# ===========================================================================
# RATE CARD MANAGEMENT (Admin + Management only)
# ===========================================================================

@login_required
def rate_card_list(request):
    denied = _require_rate_card_access(request.user)
    if denied:
        return denied

    q = request.GET.get("q", "").strip()
    afe_line_id = request.GET.get("afe_line", "")
    phase = request.GET.get("phase", "")

    qs = RateCardItem.objects.select_related("afe_line").all()
    if q:
        qs = qs.filter(Q(code__icontains=q) | Q(description__icontains=q))
    if afe_line_id:
        qs = qs.filter(afe_line_id=afe_line_id)
    if phase:
        qs = qs.filter(phase_flag=phase)

    afe_lines = AFETemplate.objects.filter(is_subtotal_row=False).order_by("order")

    return render(request, "afe/rate_cards.html", {
        "items": qs[:500],
        "q": q,
        "afe_lines": afe_lines,
        "current_afe_line": afe_line_id,
        "current_phase": phase,
        "total_count": RateCardItem.objects.count(),
    })


@login_required
def rate_card_create(request):
    denied = _require_rate_card_access(request.user)
    if denied:
        return denied

    if request.method == "POST":
        form = RateCardItemForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Rate card item ditambahkan.")
            return redirect("afe:rate_card_list")
    else:
        form = RateCardItemForm()

    return render(request, "afe/rate_card_form.html", {
        "form": form,
        "title": "Tambah Rate Card Item",
        "is_new": True,
    })


@login_required
def rate_card_edit(request, pk):
    denied = _require_rate_card_access(request.user)
    if denied:
        return denied

    item = get_object_or_404(RateCardItem, pk=pk)
    if request.method == "POST":
        form = RateCardItemForm(request.POST, instance=item)
        if form.is_valid():
            form.save()
            messages.success(request, f"Rate card '{item.code}' diperbarui.")
            return redirect("afe:rate_card_list")
    else:
        form = RateCardItemForm(instance=item)

    return render(request, "afe/rate_card_form.html", {
        "form": form,
        "item": item,
        "title": f"Edit Rate Card: {item.code}",
        "is_new": False,
    })


@login_required
@require_http_methods(["POST"])
def rate_card_delete(request, pk):
    denied = _require_rate_card_access(request.user)
    if denied:
        return denied

    item = get_object_or_404(RateCardItem, pk=pk)
    code = item.code
    item.delete()
    messages.success(request, f"Rate card '{code}' dihapus.")
    return redirect("afe:rate_card_list")


@login_required
def rate_card_upload(request):
    denied = _require_rate_card_access(request.user)
    if denied:
        return denied

    if request.method == "POST":
        form = RateCardUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES["excel_file"]
            # Validate file size (max 10MB)
            if excel_file.size > 10 * 1024 * 1024:
                messages.error(request, "Ukuran file terlalu besar (maks 10 MB).")
                return redirect("afe:rate_card_upload")
            # Validate file extension
            if not excel_file.name.lower().endswith((".xlsx", ".xls")):
                messages.error(request, "Format file harus .xlsx atau .xls.")
                return redirect("afe:rate_card_upload")
            try:
                created, updated = _process_rate_card_upload(excel_file)
                RateCardImportLog.objects.create(
                    uploaded_by=request.user,
                    file_name=excel_file.name,
                    items_created=created,
                    items_updated=updated,
                    notes="Uploaded via web UI",
                )
                messages.success(
                    request,
                    f"Import selesai: {created} item baru, {updated} item diperbarui.",
                )
            except Exception:
                import logging
                logging.getLogger("afe").exception("Rate card import failed: %s", excel_file.name)
                messages.error(request, "Terjadi kesalahan saat import file. Periksa format file Anda.")
            return redirect("afe:rate_card_list")
    else:
        form = RateCardUploadForm()

    recent_logs = RateCardImportLog.objects.all()[:10]
    return render(request, "afe/rate_card_upload.html", {
        "form": form,
        "recent_logs": recent_logs,
    })


def _process_rate_card_upload(excel_file) -> tuple[int, int]:
    """Parse D.MATE sheet from uploaded Excel and upsert RateCardItem records."""
    import openpyxl

    wb = openpyxl.load_workbook(excel_file, data_only=True, read_only=True)

    # Try to find D.MATE or E.COMP sheet
    sheet_name = None
    for candidate in ("D.MATE", "D-A_MEP", "E.COMP"):
        if candidate in wb.sheetnames:
            sheet_name = candidate
            break
    if sheet_name is None:
        raise ValueError(f"Sheet D.MATE atau E.COMP tidak ditemukan. Sheets: {wb.sheetnames}")

    ws = wb[sheet_name]
    created = 0
    updated = 0

    # Build AFETemplate lookup
    templates_by_code = {}
    for tpl in AFETemplate.objects.filter(is_subtotal_row=False):
        templates_by_code[tpl.line_code] = tpl

    if sheet_name == "E.COMP":
        # E.COMP format: col F(5)=AFE NO, col G(6)=description, col Q(16)=qty,
        # col S(18)=UoM, col T(19)=unit_price, col V(21)=material/non-material
        for row in ws.iter_rows(min_row=15, values_only=True):
            if not row or len(row) < 20:
                continue
            afe_no = row[5] if len(row) > 5 else None
            desc = row[6] if len(row) > 6 else None
            qty = row[16] if len(row) > 16 else None
            uom = row[18] if len(row) > 18 else None
            price = row[19] if len(row) > 19 else None
            mat_type = row[21] if len(row) > 21 else None

            if not desc or not price:
                continue
            desc_str = str(desc).strip()
            if not desc_str or "SUB TOTAL" in desc_str.upper():
                continue

            try:
                price_dec = Decimal(str(price).strip())
            except (InvalidOperation, ValueError):
                continue

            afe_no_str = str(int(afe_no)) if isinstance(afe_no, (int, float)) else str(afe_no or "")
            template = templates_by_code.get(afe_no_str)

            code = f"EC-{afe_no_str}-{created + updated + 1}"

            _, was_created = RateCardItem.objects.update_or_create(
                code=code[:30],
                description=desc_str[:300],
                defaults={
                    "unit_of_measure": str(uom or "")[:30],
                    "unit_price_usd": price_dec,
                    "afe_line": template,
                    "material_type": str(mat_type or "")[:30],
                    "source_sheet": sheet_name,
                },
            )
            if was_created:
                created += 1
            else:
                updated += 1

    return created, updated
