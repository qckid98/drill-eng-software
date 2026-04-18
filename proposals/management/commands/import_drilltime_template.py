"""
Import a drilling time template from the Excel workbook.

Creates a Proposal with status=TEMPLATE containing:
  - Well data from A.Proposal sheet (rows 3-27)
  - CasingSection rows from A.Proposal casing table (rows 36-43)
  - ProposalActivity rows from Tab 2.0 (section-activity assignments)
  - TubingSpec rows from A.Proposal (rows 49-51)
  - OperationalRate rows from A.Proposal (rows 46-50, cols P-W)
  - FormationMarker rows from A.Proposal (rows 46-48, cols L-M)
  - CompletionSpec from A.Proposal (rows 62-65)

Usage:
    python manage.py import_drilltime_template "path/to/DrillTime_*.xlsx"

Requires: master data already imported via import_drilltime_master.
"""
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.contrib.auth import get_user_model
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from masterdata.models import (
    DrillingActivity,
    HoleSection,
    MudType,
)
from proposals.models import (
    CasingSection,
    CompletionSpec,
    FormationMarker as ProposalFormationMarker,
    OperationalRate,
    Proposal,
    ProposalActivity,
    ProposalStatus,
    TubeLengthRange,
    TubingItem,
)
from proposals.services.calc import recalculate_proposal
from wells.models import FormationMarker, Well

User = get_user_model()


def _to_decimal(value):
    if value in (None, "", "--------", "#N/A", "#VALUE!"):
        return None
    s = str(value).strip()
    if s.startswith("="):
        return None
    try:
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return None


def _text(value):
    if value is None:
        return ""
    s = str(value).strip()
    if s in ("--------", "#N/A", "#VALUE!", "None"):
        return ""
    return s


class Command(BaseCommand):
    help = "Import a drilling time template proposal from the source Excel workbook."

    def add_arguments(self, parser):
        parser.add_argument("excel_path", type=str, help="Path to DrillTime_*.xlsx")
        parser.add_argument(
            "--dry-run", action="store_true",
            help="Parse and report but don't write to database.",
        )

    def handle(self, *args, **options):
        path = Path(options["excel_path"])
        if not path.exists():
            raise CommandError(f"File not found: {path}")

        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise CommandError("openpyxl is not installed.") from exc

        self.stdout.write(f"Loading workbook {path.name} ...")
        self.wb = load_workbook(path, data_only=True, read_only=True)

        # Need at least one user to assign as creator
        admin_user = User.objects.filter(is_superuser=True).first()
        if not admin_user:
            admin_user = User.objects.first()
        if not admin_user:
            raise CommandError(
                "No users in database. Create a superuser first: "
                "python manage.py createsuperuser"
            )

        with transaction.atomic():
            well = self._import_well()
            proposal = self._create_template_proposal(well, admin_user)
            self._import_casing_sections(proposal)
            self._import_tab2_activities(proposal)
            self._import_tubing(proposal)
            self._import_operational_rates(proposal)
            self._import_formation_markers(well, proposal)
            self._import_tube_length_ranges(proposal)
            self._import_rig_spec(proposal)
            self._import_completion_spec(proposal)

            # Recalculate totals (also triggers overlap liner calc)
            recalculate_proposal(proposal)

            if options["dry_run"]:
                transaction.set_rollback(True)
                self.stdout.write(self.style.WARNING("Dry run -- rolled back."))
            else:
                self.stdout.write(self.style.SUCCESS(
                    f"Template proposal created: PK={proposal.pk}, "
                    f"doc={proposal.doc_number}, "
                    f"sections={proposal.casing_sections.count()}, "
                    f"total_rig_days={proposal.total_rig_days}"
                ))

    def _get_cell(self, ws, row, col):
        """Get cell value from a read-only worksheet."""
        for r in ws.iter_rows(min_row=row, max_row=row, min_col=col, max_col=col, values_only=True):
            return r[0] if r else None
        return None

    # ------------------------------------------------------------------
    # A.Proposal -> Well
    # ------------------------------------------------------------------
    def _import_well(self):
        ws = self.wb["A.Proposal"]
        rows = {}
        for r in range(1, 30):
            row_data = []
            for cell_row in ws.iter_rows(min_row=r, max_row=r, values_only=True):
                row_data = list(cell_row)
            rows[r] = row_data

        def _val(row_num, col_idx):
            r = rows.get(row_num, [])
            return r[col_idx] if col_idx < len(r) else None

        well, _ = Well.objects.update_or_create(
            name=_text(_val(13, 5)) or "TEMPLATE WELL",
            location=_text(_val(14, 5)) or "TEMPLATE",
            defaults={
                "cluster": _text(_val(13, 5)),
                "field": _text(_val(15, 5)),
                "basin": _text(_val(16, 5)),
                "operator": _text(_val(4, 5)) or "PT. PERTAMINA EP",
                "contract_area": f"{_text(_val(5, 5))} / {_text(_val(6, 5))}".strip(" /"),
                "elevation_m": _to_decimal(_val(17, 5)),
                "target_formation": _text(_val(18, 5)),
                "surface_lat": _text(_val(19, 5)),
                "surface_lon": _text(_val(19, 7)),
                "target_lat": _text(_val(20, 5)),
                "target_lon": _text(_val(20, 7)),
                "kop_m": _to_decimal(_val(23, 5)),
                "inclination_deg": _to_decimal(_val(26, 5)),
                "azimuth_deg": _to_decimal(_val(27, 5)),
                "overlap_liner_7in_m": _to_decimal(_val(24, 5)),
                "overlap_liner_4in_m": _to_decimal(_val(25, 5)),
            },
        )
        self.stdout.write(f"  Well: {well}")
        return well

    # ------------------------------------------------------------------
    # Create template proposal
    # ------------------------------------------------------------------
    def _create_template_proposal(self, well, user):
        ws = self.wb["A.Proposal"]
        rows = {}
        for r in range(1, 80):
            row_data = []
            for cell_row in ws.iter_rows(min_row=r, max_row=r, values_only=True):
                row_data = list(cell_row)
            rows[r] = row_data

        def _val(row_num, col_idx):
            r = rows.get(row_num, [])
            return r[col_idx] if col_idx < len(r) else None

        from masterdata.models import RigSpec
        rig = RigSpec.objects.first()

        # Check for existing template and delete it
        Proposal.objects.filter(
            status=ProposalStatus.TEMPLATE,
            well=well,
        ).delete()

        proposal = Proposal(
            well=well,
            rig=rig,
            created_by=user,
            title=f"Template - {well.name} {well.location}",
            status=ProposalStatus.TEMPLATE,
            mob_days=_to_decimal(_val(74, 5)) or Decimal("0"),
            demob_days=_to_decimal(_val(75, 5)) or Decimal("0"),
            dollar_rate=_to_decimal(_val(60, 5)),  # row 60 col F = KURS DOLLAR
        )
        # Set spud date if available
        spud = _val(68, 5)
        if spud and hasattr(spud, "date"):
            proposal.spud_date = spud.date() if callable(getattr(spud, "date", None)) else spud
        elif spud and hasattr(spud, "year"):
            import datetime
            proposal.spud_date = datetime.date(spud.year, spud.month, spud.day)

        date_input = _val(9, 5)
        if date_input and hasattr(date_input, "date"):
            proposal.date_input = date_input.date() if callable(getattr(date_input, "date", None)) else date_input
        elif date_input and hasattr(date_input, "year"):
            import datetime
            proposal.date_input = datetime.date(date_input.year, date_input.month, date_input.day)

        proposal.save()
        self.stdout.write(f"  Proposal template: {proposal.doc_number}")
        return proposal

    # ------------------------------------------------------------------
    # A.Proposal rows 36-43 -> CasingSection
    # ------------------------------------------------------------------
    def _import_casing_sections(self, proposal):
        ws = self.wb["A.Proposal"]
        count = 0
        for r in range(36, 44):
            row_data = []
            for cell_row in ws.iter_rows(min_row=r, max_row=r, values_only=True):
                row_data = list(cell_row)

            def _v(idx):
                return row_data[idx] if idx < len(row_data) else None

            hole_size = _to_decimal(_v(5))  # col F = Open Hole inch
            od_csg = _to_decimal(_v(6))     # col G = OD CSG
            if hole_size is None:
                continue

            # Find or skip HoleSection
            try:
                hole_section = HoleSection.objects.get(size_inch=hole_size)
            except HoleSection.DoesNotExist:
                # Create it
                label = f'{hole_size.normalize():f}"'
                order_value = max(0, int((Decimal("100") - hole_size) * 10))
                hole_section = HoleSection.objects.create(
                    size_inch=hole_size, label=label, order=order_value,
                )

            depth = _to_decimal(_v(9)) or Decimal("0")  # col J = DEPTH
            mud_name = _text(_v(11))                      # col L = MUD TYPE
            mud_type = None
            if mud_name:
                mud_type, _ = MudType.objects.get_or_create(name=mud_name)

            is_completion = _text(_v(5)).upper() == "COMPLETION"

            section = CasingSection.objects.create(
                proposal=proposal,
                order=count,
                hole_section=hole_section,
                od_csg=od_csg,
                id_csg=_to_decimal(_v(8)),          # col I = ID CSG
                weight_lbs_ft=_to_decimal(_v(7)),   # col H = WEIGHT
                depth_m=depth,
                top_of_liner_m=_to_decimal(_v(10)), # col K = TOL (mMD)
                mud_type=mud_type,
                is_completion=is_completion,
                sg_from=_to_decimal(_v(13)),         # col N = SG FR
                sg_to=_to_decimal(_v(14)),           # col O = SG TO
                casing_type=_text(_v(19)),           # col T = TYPE (K-55)
                pounder=_to_decimal(_v(20)),         # col U = POUNDER
                thread=_text(_v(21)),                # col V = THREAD
                range_spec=_text(_v(22)),            # col W = RANGE
                casing_title=_text(_v(23)),          # col X = TITLE
            )
            count += 1

        self.stdout.write(f"  CasingSections created: {count}")

    # ------------------------------------------------------------------
    # Tab 2.0 -> ProposalActivity (mapped to CasingSections)
    # ------------------------------------------------------------------
    def _import_tab2_activities(self, proposal):
        if "Tab 2.0" not in self.wb.sheetnames:
            self.stdout.write(self.style.WARNING("  Tab 2.0 missing, skipping activities."))
            return

        ws = self.wb["Tab 2.0"]
        sections = list(proposal.casing_sections.order_by("order"))
        if not sections:
            self.stdout.write(self.style.WARNING("  No casing sections, skipping Tab 2.0."))
            return

        # Build a lookup: hole_size -> CasingSection
        # Tab 2.0 col C (idx 2) has hole section info like "Pre Spud", "26", "17.5", etc.
        # We need to map section types to our CasingSection records
        section_by_size = {}
        for s in sections:
            size_key = str(s.hole_section.size_inch.normalize())
            section_by_size[size_key] = s

        # Also map special sections
        # "Pre Spud" -> first section (largest hole or order=0)
        # "Rig Release" -> last section
        # "Completion" -> section with is_completion=True
        first_section = sections[0] if sections else None
        last_section = sections[-1] if sections else None
        completion_section = next((s for s in sections if s.is_completion), last_section)

        activity_count = 0
        current_section = first_section
        act_order = 0

        for row in ws.iter_rows(min_row=9, max_row=ws.max_row, values_only=True):
            if not row:
                continue

            at_num = row[1] if len(row) > 1 else None       # col B = AT number
            hole_sz = _text(row[2]) if len(row) > 2 else ""  # col C = hole section
            sect_type = _text(row[3]) if len(row) > 3 else ""  # col D = section type
            cat1 = _text(row[4]) if len(row) > 4 else ""     # col E = category 1
            cat2 = _text(row[5]) if len(row) > 5 else ""     # col F = category 2
            code = _text(row[16]) if len(row) > 16 else ""   # col Q = code
            hrs = _to_decimal(row[17]) if len(row) > 17 else None  # col R = hours
            depth = _to_decimal(row[19]) if len(row) > 19 else None  # col T = depth

            # Skip separator/empty rows
            if not cat2 and not cat1:
                continue
            if "---" in cat2 or "---" in cat1:
                continue

            # New section group (AT number present)
            if at_num is not None and at_num != "":
                act_order = 0
                # Determine which CasingSection this maps to
                hole_upper = hole_sz.upper().strip()
                if "PRE" in hole_upper or "SPUD" in hole_upper:
                    current_section = first_section
                elif "RIG RE" in hole_upper or "RELEASE" in hole_upper:
                    current_section = last_section
                elif "COMPLE" in hole_upper:
                    current_section = completion_section
                else:
                    # Try to match by hole size number
                    size_val = _to_decimal(hole_sz)
                    if size_val is not None:
                        size_key = str(size_val.normalize())
                        if size_key in section_by_size:
                            current_section = section_by_size[size_key]

                # Update section_type on the CasingSection
                if current_section and sect_type:
                    current_section.section_type = sect_type[:200]
                    current_section.save(update_fields=["section_type"])

            if current_section is None:
                continue

            # Skip rows without valid hours or with #N/A
            if hrs is None or hrs <= 0:
                continue

            # Find matching DrillingActivity from master data
            activity = None
            if cat2:
                # Try exact match on description
                candidates = DrillingActivity.objects.filter(
                    description__icontains=cat2[:80]
                )
                if candidates.exists():
                    activity = candidates.first()

            if activity is None and code:
                # Try match by code
                candidates = DrillingActivity.objects.filter(code=code[:20])
                if candidates.exists():
                    activity = candidates.first()

            if activity is None:
                # Create a placeholder activity
                from masterdata.models import ActivityCategoryL1, ActivityCategoryL2
                l1, _ = ActivityCategoryL1.objects.get_or_create(
                    name=cat1[:150] or "UNCATEGORIZED",
                    defaults={"order": 999},
                )
                l2, _ = ActivityCategoryL2.objects.get_or_create(
                    parent=l1,
                    name=cat1[:250] or "UNCATEGORIZED",
                    defaults={"order": 0},
                )
                activity, _ = DrillingActivity.objects.get_or_create(
                    category_l2=l2,
                    code=code[:20] or "---",
                    description=cat2[:500] or "Unknown activity",
                    defaults={
                        "default_hours": hrs or Decimal("0"),
                        "phase_type": guess_phase_type_simple(cat1),
                    },
                )

            ProposalActivity.objects.create(
                casing_section=current_section,
                activity=activity,
                order=act_order,
                hours_override=hrs,
                notes=f"Imported from Tab 2.0",
            )
            act_order += 1
            activity_count += 1

        self.stdout.write(f"  ProposalActivities created: {activity_count}")

    # ------------------------------------------------------------------
    # A.Proposal rows 49-51 -> TubingItem
    # ------------------------------------------------------------------
    def _import_tubing(self, proposal):
        ws = self.wb["A.Proposal"]
        count = 0
        for r in range(49, 52):
            row_data = []
            for cell_row in ws.iter_rows(min_row=r, max_row=r, values_only=True):
                row_data = list(cell_row)

            def _v(idx):
                return row_data[idx] if idx < len(row_data) else None

            od = _to_decimal(_v(5))
            if od is None:
                continue

            TubingItem.objects.create(
                proposal=proposal,
                order=count,
                od_inch=od,
                id_inch=_to_decimal(_v(6)),
                weight_lbs_ft=_to_decimal(_v(7)),
                avg_length_m=_to_decimal(_v(8)),
                depth_md=_to_decimal(_v(9)),
            )
            count += 1
        self.stdout.write(f"  TubingItems created: {count}")

    # ------------------------------------------------------------------
    # A.Proposal rows 46-50 -> OperationalRate
    # ------------------------------------------------------------------
    def _import_operational_rates(self, proposal):
        ws = self.wb["A.Proposal"]
        rates_data = [
            # (row, name_col, value_col, unit_col)
            (46, 19, 22, 23),  # LENGTH PER STAND
            (47, 19, 22, 23),  # RATE FOR RIH/POOH BHA & TBG
            (48, 19, 22, 23),  # RATE FOR RUNNING CASING
            (49, 19, 22, 23),  # RATE FOR STAND-UP
            (50, 19, 22, 23),  # LENGTH PER JOINT
        ]
        count = 0
        for idx, (r, name_col, val_col, unit_col) in enumerate(rates_data):
            row_data = []
            for cell_row in ws.iter_rows(min_row=r, max_row=r, values_only=True):
                row_data = list(cell_row)

            def _v(c):
                return row_data[c] if c < len(row_data) else None

            name = _text(_v(name_col))
            value = _to_decimal(_v(val_col))
            unit = _text(_v(unit_col))

            if not name or value is None:
                continue

            OperationalRate.objects.create(
                proposal=proposal,
                rate_name=name[:120],
                value=value,
                unit=unit[:30],
                order=idx,
            )
            count += 1
        self.stdout.write(f"  OperationalRates created: {count}")

    # ------------------------------------------------------------------
    # A.Proposal rows 46-52 col L-M -> FormationMarker (Well + Proposal)
    # ------------------------------------------------------------------
    def _import_formation_markers(self, well, proposal):
        ws = self.wb["A.Proposal"]
        count = 0
        for r in range(46, 53):
            row_data = []
            for cell_row in ws.iter_rows(min_row=r, max_row=r, values_only=True):
                row_data = list(cell_row)

            def _v(idx):
                return row_data[idx] if idx < len(row_data) else None

            name = _text(_v(11))   # col L = formation name
            depth = _to_decimal(_v(12))  # col M = depth

            if not name or depth is None:
                continue

            # Well-level formation marker
            FormationMarker.objects.update_or_create(
                well=well,
                name=name[:120],
                defaults={"depth_m": depth, "order": count},
            )
            # Proposal-level formation marker
            ProposalFormationMarker.objects.create(
                proposal=proposal,
                name=name[:120],
                depth_md=depth,
                order=count,
            )
            count += 1
        self.stdout.write(f"  FormationMarkers created: {count}")

    # ------------------------------------------------------------------
    # A.Proposal rows 46-49 col P-R -> TubeLengthRange
    # ------------------------------------------------------------------
    def _import_tube_length_ranges(self, proposal):
        ws = self.wb["A.Proposal"]
        count = 0
        for r in range(46, 50):  # rows 46-49: R-1, R-2, R-3, SP
            row_data = []
            for cell_row in ws.iter_rows(min_row=r, max_row=r, values_only=True):
                row_data = list(cell_row)

            def _v(idx):
                return row_data[idx] if idx < len(row_data) else None

            label = _text(_v(15))          # col P = label (R-1, R-2, R-3, SP)
            avg_length = _to_decimal(_v(16))  # col Q = avg length (m)

            if not label or avg_length is None:
                continue

            TubeLengthRange.objects.create(
                proposal=proposal,
                label=label[:10],
                avg_length_m=avg_length,
            )
            count += 1
        self.stdout.write(f"  TubeLengthRanges created: {count}")

    # ------------------------------------------------------------------
    # A.Proposal rows 54-59 -> RigSpec (link to Proposal)
    # ------------------------------------------------------------------
    def _import_rig_spec(self, proposal):
        ws = self.wb["A.Proposal"]
        rows = {}
        for r in range(53, 61):
            row_data = []
            for cell_row in ws.iter_rows(min_row=r, max_row=r, values_only=True):
                row_data = list(cell_row)
            rows[r] = row_data

        def _val(row_num, col_idx):
            r = rows.get(row_num, [])
            return r[col_idx] if col_idx < len(r) else None

        platform_name = _text(_val(54, 5))  # row 54 col F = PLATFORM
        if not platform_name:
            self.stdout.write("  RigSpec: no platform name, skipping.")
            return

        from masterdata.models import RigSpec
        rig, created = RigSpec.objects.update_or_create(
            platform_name=platform_name[:100],
            defaults={
                "floor_height_m": _to_decimal(_val(55, 5)),  # row 55 = TINGGI RIG FLOOR
                "horsepower": int(_to_decimal(_val(56, 5)) or 0) or None,  # row 56 = HP
                "status": _text(_val(57, 5)),                 # row 57 = STATUS RIG
                "capacity": _text(_val(59, 5)),               # row 59 = KAPASITAS RIG
            },
        )
        # Link rig to proposal
        proposal.rig = rig
        proposal.save(update_fields=["rig"])
        self.stdout.write(f"  RigSpec: {rig.platform_name} ({'created' if created else 'updated'})")

    # ------------------------------------------------------------------
    # A.Proposal rows 62-65 -> CompletionSpec
    # Row 62: Jenis Garam (col I=8) = CaCl2
    # Row 63: SG CF (col I=8) = 1.2
    # Row 64: Yield per SG (col I=8) = 101.8
    # Row 65: Packaging (col I=8) = 50 kg/sax
    # ------------------------------------------------------------------
    def _import_completion_spec(self, proposal):
        ws = self.wb["A.Proposal"]
        rows = {}
        for r in range(62, 70):
            row_data = []
            for cell_row in ws.iter_rows(min_row=r, max_row=r, values_only=True):
                row_data = list(cell_row)
            rows[r] = row_data

        def _val(row_num, col_idx):
            r = rows.get(row_num, [])
            return r[col_idx] if col_idx < len(r) else None

        CompletionSpec.objects.update_or_create(
            proposal=proposal,
            defaults={
                "salt_type": _text(_val(62, 8)) or "CaCl2",   # col I row 62
                "sg": _to_decimal(_val(63, 8)),                 # col I row 63
                "yield_value": _to_decimal(_val(64, 8)),        # col I row 64
                "packaging_kg_per_sax": _to_decimal(_val(65, 8)),  # col I row 65
            },
        )
        self.stdout.write("  CompletionSpec created")


def guess_phase_type_simple(cat1_name: str) -> str:
    """Simple phase type guesser for Tab 2.0 categories."""
    from masterdata.models import PhaseType
    upper = (cat1_name or "").upper()
    completion_kw = ("COMPLETION", "PERFORATION", "PACKER", "TUBING")
    non_drill_kw = ("RIG UP", "RIG DOWN", "MOVING", "MOVE", "ENDURANCE",
                    "WAIT", "STANDBY", "NIPPLE", "BOP", "WELLHEAD")
    for kw in completion_kw:
        if kw in upper:
            return PhaseType.COMPLETION
    for kw in non_drill_kw:
        if kw in upper:
            return PhaseType.NON_DRILLING
    return PhaseType.DRILLING
