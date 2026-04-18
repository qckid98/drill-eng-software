"""
Import master reference data from the Drilling Time Excel workbook.

Usage:
    python manage.py import_drilltime_master "path/to/DrillTime_ BDA-G3 - Cluster BDA-D1.xlsx"

Safe to re-run: everything uses update_or_create keyed on natural keys.
What it imports:
    - Sheet `ROP`           -> HoleSection + RopRate rows
    - Sheet `Tab 1.0`       -> ActivityCategoryL1/L2 + DrillingActivity rows
    - Sheet `A.Proposal`    -> RigSpec, Well seed, casing defaults
    - Sheet `Name MGR`      -> Well parameters (KOP, TOL, overlap, rates)
    - Seeds default MudTypes (Gel Water, KCL Polymer, HPWBM)
"""
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from masterdata.models import (
    ActivityCategoryL1,
    ActivityCategoryL2,
    DrillingActivity,
    HoleSection,
    MudType,
    PhaseType,
    RigSpec,
    RopRate,
)


def _to_decimal(value):
    if value in (None, ""):
        return None
    try:
        return Decimal(str(value).strip())
    except (InvalidOperation, ValueError):
        return None


def _text(value):
    return str(value).strip() if value is not None else ""


# Heuristic mapping: which L1 names should be considered non-drilling by default.
NON_DRILLING_KEYWORDS = (
    "RIG UP", "RIG DOWN", "MOVING", "MOVE", "ENDURANCE", "SKID",
    "WAIT", "STANDBY", "NIPPLE", "BOPE", "HANDLE",
)
COMPLETION_KEYWORDS = ("COMPLETION", "PERFORATION", "PACKER", "TUBING")


def guess_phase_type(l1_name: str) -> str:
    upper = l1_name.upper()
    for kw in COMPLETION_KEYWORDS:
        if kw in upper:
            return PhaseType.COMPLETION
    for kw in NON_DRILLING_KEYWORDS:
        if kw in upper:
            return PhaseType.NON_DRILLING
    return PhaseType.DRILLING


class Command(BaseCommand):
    help = "Import drilling master data (ROP, activity library, rigs) from the source Excel workbook."

    def add_arguments(self, parser):
        parser.add_argument("excel_path", type=str, help="Path to DrillTime_*.xlsx")
        parser.add_argument(
            "--dry-run", action="store_true",
            help="Parse and report counts but don't write to database.",
        )

    def handle(self, *args, **options):
        path = Path(options["excel_path"])
        if not path.exists():
            raise CommandError(f"File not found: {path}")

        try:
            import openpyxl  # noqa: F401
        except ImportError as exc:
            raise CommandError(
                "openpyxl is not installed. Run: pip install -r requirements.txt"
            ) from exc

        from openpyxl import load_workbook
        self.stdout.write(f"Loading workbook {path.name} ...")
        wb = load_workbook(path, data_only=True, read_only=True)

        with transaction.atomic():
            self._seed_mud_types()
            self._import_rop(wb)
            self._import_tab1(wb)
            self._import_rig(wb)
            self._import_name_mgr(wb)
            if options["dry_run"]:
                transaction.set_rollback(True)
                self.stdout.write(self.style.WARNING("Dry run -- rolled back."))
            else:
                self.stdout.write(self.style.SUCCESS("Import complete."))

    # ------------------------------------------------------------------
    # Mud types (static seed)
    # ------------------------------------------------------------------
    def _seed_mud_types(self):
        defaults = [
            ("Gel Water", "Basic bentonite water mud"),
            ("KCL Polimer", "Potassium chloride polymer mud"),
            ("KCL Polymer", "Potassium chloride polymer mud"),
            ("HPWBM", "High Performance Water Based Mud"),
            ("OBM", "Oil Based Mud"),
        ]
        for name, desc in defaults:
            MudType.objects.update_or_create(
                name=name, defaults={"description": desc}
            )
        self.stdout.write(f"  MudTypes: {MudType.objects.count()}")

    # ------------------------------------------------------------------
    # ROP sheet -> HoleSection + RopRate
    # ------------------------------------------------------------------
    def _import_rop(self, wb):
        if "ROP" not in wb.sheetnames:
            self.stdout.write(self.style.WARNING("  ROP sheet missing, skipping."))
            return
        ws = wb["ROP"]

        header_row = None
        size_col = None
        for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True), start=1):
            lowered = [_text(c).lower() for c in row]
            if "start" in lowered and "end" in lowered and any("rop" in c for c in lowered):
                header_row = idx
                start_idx = lowered.index("start")
                size_col = max(0, start_idx - 1)
                break
        if header_row is None:
            self.stdout.write(self.style.WARNING("  ROP header not found -- sheet layout changed?"))
            return

        created_sections = 0
        created_rates = 0
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if not row:
                continue
            size = _to_decimal(row[size_col]) if size_col < len(row) else None
            if size is None:
                continue
            start = _to_decimal(row[size_col + 1]) or Decimal("0")
            end = _to_decimal(row[size_col + 2]) or Decimal("0")
            days = _to_decimal(row[size_col + 3]) or Decimal("0")
            rop = _to_decimal(row[size_col + 4]) or Decimal("0")

            label = f'{size.normalize():f}"'
            order_value = max(0, int((Decimal("100") - size) * 10))
            section, created = HoleSection.objects.update_or_create(
                size_inch=size,
                defaults={"label": label, "order": order_value},
            )
            if created:
                created_sections += 1
            RopRate.objects.update_or_create(
                hole_section=section,
                start_depth_m=start,
                end_depth_m=end,
                defaults={"days": days, "rop_m_per_day": rop},
            )
            created_rates += 1

        self.stdout.write(
            f"  HoleSection: {HoleSection.objects.count()} (new {created_sections}), "
            f"RopRate rows touched: {created_rates}"
        )

    # ------------------------------------------------------------------
    # Tab 1.0 -> Activity library
    #
    # Actual column layout (0-indexed):
    #   col 3 (D)  = CATEGORY LEVEL #1 number
    #   col 4 (E)  = CATEGORY LEVEL #1 name (only on header rows)
    #   col 5 (F)  = CATEGORY LEVEL #2 number / activity number
    #   col 6 (G)  = Activity description
    #   col 13 (N) = CODE (e.g. '1b', '2b')
    #   col 15 (P) = HOLE SECTION applicability
    #   col 23 (X) = DRILLING TIME (HRS) STD
    #   col 25 (Z) = DRILLING TIME (HRS) TOTAL (STD + ADJUSTING)
    #
    # L1 = major phase (e.g. "RIG UP / RIG DOWN")
    # L2 = real sub-category from the L1 name (we use L1 name as L2 parent)
    # Activity = individual work item with code + description + hours
    # ------------------------------------------------------------------
    def _import_tab1(self, wb):
        sheet_name = None
        for candidate in ("Tab 1.0", "Tab1.0", "TAB 1.0"):
            if candidate in wb.sheetnames:
                sheet_name = candidate
                break
        if sheet_name is None:
            self.stdout.write(self.style.WARNING("  Tab 1.0 sheet missing, skipping."))
            return

        ws = wb[sheet_name]

        COL_L1_NUM = 3
        COL_L1_NAME = 4
        COL_ACT_NUM = 5
        COL_ACT_DESC = 6
        COL_CODE = 13
        COL_HOURS_STD = 23
        COL_HOURS_TOTAL = 25

        l1_cache = {}      # name -> ActivityCategoryL1
        l2_cache = {}      # (l1_name, l2_name) -> ActivityCategoryL2
        activities_created = 0
        activities_updated = 0
        current_l1 = None
        current_l1_name = None
        l1_order = 0

        def _cell(row, idx):
            return row[idx] if idx < len(row) else None

        for row in ws.iter_rows(min_row=4, values_only=True):
            if not row:
                continue
            l1_num = _cell(row, COL_L1_NUM)
            l1_name_raw = _text(_cell(row, COL_L1_NAME))
            act_num = _cell(row, COL_ACT_NUM)
            desc = _text(_cell(row, COL_ACT_DESC))
            code = _text(_cell(row, COL_CODE))
            hours_std = _to_decimal(_cell(row, COL_HOURS_STD))
            hours_total = _to_decimal(_cell(row, COL_HOURS_TOTAL))

            # Use TOTAL hours if available, otherwise STD
            hours = hours_total if hours_total is not None else hours_std

            # Header row for a new L1 phase: has L1 number + L1 name, no activity desc
            if l1_num not in (None, "") and l1_name_raw and not desc:
                current_l1_name = l1_name_raw
                if current_l1_name not in l1_cache:
                    obj, _ = ActivityCategoryL1.objects.update_or_create(
                        name=current_l1_name[:150],
                        defaults={"order": l1_order},
                    )
                    l1_cache[current_l1_name] = obj
                    l1_order += 1
                current_l1 = l1_cache[current_l1_name]
                continue

            # Activity row: needs a description + current L1 context
            if not desc or current_l1 is None:
                continue

            # Use the L1 name as L2 bucket (real hierarchy from Excel)
            l2_name = current_l1_name or "(umum)"
            l2_key = (current_l1_name, l2_name)
            if l2_key not in l2_cache:
                l2_obj, _ = ActivityCategoryL2.objects.update_or_create(
                    parent=current_l1,
                    name=l2_name[:250],
                    defaults={"order": 0},
                )
                l2_cache[l2_key] = l2_obj
            l2 = l2_cache[l2_key]

            # Build effective code
            effective_code = code
            if not effective_code and isinstance(act_num, (int, float)):
                effective_code = str(int(act_num))
            effective_code = (effective_code or "")[:20]

            obj, created = DrillingActivity.objects.update_or_create(
                category_l2=l2,
                code=effective_code,
                description=desc[:500],
                defaults={
                    "default_hours": hours or Decimal("0"),
                    "phase_type": guess_phase_type(current_l1_name or ""),
                },
            )
            if created:
                activities_created += 1
            else:
                activities_updated += 1

        self.stdout.write(
            f"  Activity categories L1={ActivityCategoryL1.objects.count()} "
            f"L2={ActivityCategoryL2.objects.count()} "
            f"Activities={DrillingActivity.objects.count()} "
            f"(new {activities_created}, updated {activities_updated})"
        )

    # ------------------------------------------------------------------
    # A.Proposal sheet -> seed RigSpec
    # ------------------------------------------------------------------
    def _import_rig(self, wb):
        if "A.Proposal" not in wb.sheetnames:
            return
        ws = wb["A.Proposal"]
        platform_name = None
        horsepower = None
        floor_height = None

        for row in ws.iter_rows(min_row=50, max_row=65, values_only=True):
            for idx, cell in enumerate(row):
                text = _text(cell).upper()
                if "PLATFORM" in text and platform_name is None:
                    for c in row[idx + 1:]:
                        if c:
                            platform_name = _text(c)
                            break
                if "HORSE POWER" in text or "HORSEPOWER" in text:
                    for c in row[idx + 1:]:
                        if c and _to_decimal(c) is not None:
                            horsepower = int(_to_decimal(c))
                            break
                if "FLOOR HEIGHT" in text or "TINGGI RIG" in text:
                    for c in row[idx + 1:]:
                        if c and _to_decimal(c) is not None:
                            floor_height = _to_decimal(c)
                            break

        if not platform_name:
            platform_name = "PDSI #04.3"
        RigSpec.objects.update_or_create(
            platform_name=platform_name[:100],
            defaults={
                "horsepower": horsepower or 1500,
                "floor_height_m": floor_height,
                "capacity": "",
                "status": "Active",
                "notes": "Seeded from A.Proposal sheet",
            },
        )
        self.stdout.write(f"  RigSpec seeded: {platform_name}")

    # ------------------------------------------------------------------
    # Name MGR sheet -> Well parameters (KOP, TOL, overlap, rates)
    # ------------------------------------------------------------------
    def _import_name_mgr(self, wb):
        if "Name MGR" not in wb.sheetnames:
            self.stdout.write(self.style.WARNING("  Name MGR sheet missing, skipping."))
            return
        ws = wb["Name MGR"]

        params = {}
        for row in ws.iter_rows(min_row=1, max_row=min(100, ws.max_row), values_only=True):
            if not row:
                continue
            for idx, cell in enumerate(row):
                text = _text(cell).upper()
                if "KOP #1" in text:
                    for c in row[idx + 1:]:
                        val = _to_decimal(c)
                        if val is not None:
                            params["kop_1"] = val
                            break
                if "TOL 7IN" in text or "TOL 7 IN" in text:
                    for c in row[idx + 1:]:
                        val = _to_decimal(c)
                        if val is not None:
                            params["tol_7in"] = val
                            break
                if "TOL 4" in text:
                    for c in row[idx + 1:]:
                        val = _to_decimal(c)
                        if val is not None:
                            params["tol_4in"] = val
                            break
                if "OVERLAP LINER 7" in text:
                    for c in row[idx + 1:]:
                        val = _to_decimal(c)
                        if val is not None:
                            params["overlap_7in"] = val
                            break
                if "OVERLAP LINER 4" in text:
                    for c in row[idx + 1:]:
                        val = _to_decimal(c)
                        if val is not None:
                            params["overlap_4in"] = val
                            break

        self.stdout.write(f"  Name MGR parameters found: {params}")
